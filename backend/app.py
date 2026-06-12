from flask import Flask, request, jsonify, send_from_directory, Response
from flask_cors import CORS
import os
import re
import hashlib
import time as time_module
import logging
import requests as http_requests
from dotenv import load_dotenv
from datetime import datetime, timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

# Load environment variables FIRST
load_dotenv()

from database import (
    init_db, get_all_players, get_player_by_fid,
    save_player, delete_player, calculate_points, get_db,
    get_research_day, get_show_fire_crystals, set_setting, get_setting,
    get_time_preference_counts,
    get_time_slot_offset, set_time_slot_offset,
)
from slots import (
    VALID_SLOT_OFFSETS, slot_ids, matching_slots_for_hour, display_slot_id,
    slot_id_to_index, slot_index_to_id, slot_mapping,
)
import json

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

STATIC_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static')

app = Flask(__name__, static_folder=None)
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'dev-secret-key')
CORS(app)

# Admin credentials
ADMIN_PASSWORD = os.getenv('ADMIN_PASSWORD', 'admin123')
MINISTER_PASSWORD = os.getenv('MINISTER_PASSWORD', 'minister123')

# Salt used to sign requests to the public WOS player-lookup API at
# wos-giftcode-api.centurygame.com. Not a real secret: it's a constant
# extracted from the public gift-code site's JavaScript, shared across
# every WOS community tool. Override via env var only if upstream rotates.
WOS_API_SECRET = os.getenv('WOS_API_SECRET', 'tB87#kPtkxqOS2')

# Valid auth tokens
VALID_TOKENS = {'admin-token', 'minister-token'}


def check_admin_auth():
    """Validate admin authentication. Returns None if valid, or error response tuple if invalid."""
    auth_header = request.headers.get('Authorization')
    if auth_header and auth_header.startswith('Bearer '):
        auth_header = auth_header[len('Bearer '):]
    if not auth_header or auth_header not in VALID_TOKENS:
        return jsonify({'error': 'Unauthorized'}), 401
    return None

# Initialize database (registers teardown handler)
init_db(app)

# Health check endpoint
@app.route('/health', methods=['GET'])
def health_check():
    return jsonify({'status': 'healthy'}), 200

# Serve React app - handles SPA routing for all non-API paths.
#
# The frontend bundle is built with Vite's `base: './'` (see frontend/vite.config.ts),
# which makes asset references relative. To make those relative URLs resolve under
# the configured URL sub-path (set via URL_PREFIX), we splice a <base href> tag into
# index.html at request time. The same prefix is exposed as <meta name="app-base">
# for JS-side consumers (axios baseURL, React Router basename — see frontend/src/utils/appBase.ts).
_HEAD_RE = re.compile(r'(<head[^>]*>)', re.IGNORECASE)
_APP_BASE_MARKER = '<meta name="app-base"'
_INDEX_HTML_CACHE: dict = {}


def inject_base_tag(html: str, prefix: str) -> str:
    """Splice <base href> + <meta name="app-base"> into the <head> of an HTML document.

    `prefix` is the URL sub-path the SPA is hosted at (e.g. "" for root, "/ministry"
    for a sub-path). The injected <base href> always ends with "/" so relative URLs
    resolve correctly; the meta tag's content is the normalized prefix without a
    trailing slash (or "/" for root) which the frontend reads via getAppBase().

    Idempotent: if the marker meta tag is already present, returns the input unchanged.
    If no <head> tag is found, returns the input unchanged (malformed HTML).
    """
    if _APP_BASE_MARKER in html:
        return html
    normalized = prefix.rstrip('/')
    base_href = (normalized + '/') if normalized else '/'
    meta_content = normalized if normalized else '/'
    tags = f'<base href="{base_href}"><meta name="app-base" content="{meta_content}">'
    match = _HEAD_RE.search(html)
    if not match:
        return html
    return html[:match.end()] + tags + html[match.end():]


@app.route('/', defaults={'path': ''})
@app.route('/<path:path>')
def serve(path):
    if path:
        static_target = os.path.join(STATIC_DIR, path)
        if os.path.isfile(static_target):
            return send_from_directory(STATIC_DIR, path)
    index_path = os.path.join(STATIC_DIR, 'index.html')
    if not os.path.isfile(index_path):
        return Response('Not Found', status=404)
    prefix = request.script_root  # '' when no dispatcher, '/ministry' when prefixed
    cached = _INDEX_HTML_CACHE.get(prefix)
    if cached is None:
        with open(index_path, encoding='utf-8') as f:
            cached = inject_base_tag(f.read(), prefix)
        _INDEX_HTML_CACHE[prefix] = cached
    return Response(cached, mimetype='text/html')

# API Routes

@app.route('/api/player/submit', methods=['POST'])
def submit_player():
    """Submit or update player information."""
    try:
        data = request.json

        # Check if applications are closed (block new submissions only)
        closing_time = get_setting('application_closing_time', '')
        if closing_time:
            try:
                dt = datetime.fromisoformat(closing_time.replace('Z', '+00:00'))
                fid_str = str(data.get('fid', '')).strip()
                existing = get_player_by_fid(fid_str) if fid_str else None
                now_utc = datetime.now(timezone.utc)
                if not existing and now_utc >= dt:
                    logger.info(f"Blocked new submission from FID {fid_str} - applications closed")
                    return jsonify({'error': 'Applications are closed', 'code': 'APPLICATIONS_CLOSED'}), 403
            except ValueError:
                logger.warning(f"Invalid closing_time format in settings: {closing_time}")

        # Validate required fields
        required_fields = ['fid', 'game_name', 'alliance']
        for field in required_fields:
            if field not in data or not str(data[field]).strip():
                return jsonify({'error': f'Missing required field: {field}'}), 400

        # Extract time slots (supports both legacy list and per-day dict)
        time_slots_by_day = data.pop('time_slots_by_day', None)
        time_slots = data.pop('time_slots', [])
        if time_slots_by_day:
            time_slots = time_slots_by_day  # pass dict to save_player

        # Set defaults and validate numeric fields
        numeric_fields = [
            'construction_speedups_days', 'research_speedups_days',
            'troop_training_speedups_days', 'general_speedups_days',
            'fire_crystals', 'refined_fire_crystals', 'fire_crystal_shards'
        ]
        for field in numeric_fields:
            if field not in data:
                data[field] = 0
            else:
                try:
                    data[field] = float(data[field]) if '.' in str(data[field]) or 'speedups' in field else int(data[field])
                except (ValueError, TypeError):
                    return jsonify({'error': f'Invalid value for {field}'}), 400
                if data[field] < 0:
                    return jsonify({'error': f'{field} cannot be negative'}), 400
                if data[field] > 99999:
                    return jsonify({'error': f'{field} exceeds maximum allowed value (99999)'}), 400

        # Save player
        player_id = save_player(data, time_slots)
        logger.info(f"Player submitted/updated: FID={data.get('fid')}, name={data.get('game_name')}, id={player_id}")

        return jsonify({
            'success': True,
            'player_id': player_id,
            'message': 'Player information saved successfully'
        }), 200

    except Exception as e:
        logger.error(f"Error in submit_player: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500

@app.route('/api/player/check-duplicate', methods=['POST'])
def check_duplicate():
    """Check if a player with the given FID or game name already exists."""
    try:
        data = request.json
        fid = data.get('fid', '').strip()
        game_name = data.get('game_name', '').strip()

        db = get_db()
        cursor = db.cursor()
        result = {'fid_exists': False, 'name_exists': False}

        if fid:
            cursor.execute('SELECT id FROM players WHERE fid = ?', (fid,))
            if cursor.fetchone():
                result['fid_exists'] = True

        if game_name:
            cursor.execute('SELECT id FROM players WHERE LOWER(game_name) = LOWER(?)', (game_name,))
            if cursor.fetchone():
                result['name_exists'] = True

        return jsonify(result), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/player/<fid>', methods=['GET'])
def get_player(fid):
    """Get player information by FID."""
    try:
        player = get_player_by_fid(fid)
        if player:
            return jsonify(player), 200
        return jsonify({'error': 'Player not found'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/player/wos-lookup', methods=['POST'])
def wos_lookup():
    """Look up player info from the WOS game API."""
    try:
        data = request.json
        fid = data.get('fid', '').strip()

        if not fid:
            return jsonify({'error': 'FID is required'}), 400

        # Build signed request for WOS API
        secret = WOS_API_SECRET
        ts = str(int(time_module.time() * 1e9))
        form_data = f'fid={fid}&time={ts}'
        sign = hashlib.md5((form_data + secret).encode()).hexdigest()
        body = f'sign={sign}&{form_data}'

        response = http_requests.post(
            'https://wos-giftcode-api.centurygame.com/api/player',
            data=body,
            headers={
                'Content-Type': 'application/x-www-form-urlencoded',
                'Accept': 'application/json',
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                'Origin': 'https://wos-giftcode.centurygame.com',
                'Referer': 'https://wos-giftcode.centurygame.com/',
            },
            timeout=10
        )

        result = response.json()

        if result.get('code') != 0:
            return jsonify({'error': 'Player not found in WOS'}), 404

        wos_data = result['data']
        return jsonify({
            'success': True,
            'fid': str(wos_data['fid']),
            'nickname': wos_data.get('nickname', ''),
            'kid': wos_data.get('kid'),
            'stove_lv': wos_data.get('stove_lv'),
            'stove_lv_content': wos_data.get('stove_lv_content', ''),
            'avatar_image': wos_data.get('avatar_image', ''),
        }), 200

    except http_requests.exceptions.Timeout:
        return jsonify({'error': 'WOS API timed out'}), 504
    except http_requests.exceptions.RequestException as e:
        return jsonify({'error': f'Failed to reach WOS API: {str(e)}'}), 502
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/settings/research-day', methods=['GET'])
def get_research_day_setting():
    """Get the current research day setting (public endpoint)."""
    return jsonify({'research_day': get_research_day()}), 200

@app.route('/api/admin/settings/research-day', methods=['PUT'])
def set_research_day_setting():
    """Set the research day to 'tuesday' or 'friday'."""
    try:
        auth_error = check_admin_auth()
        if auth_error:
            return auth_error

        data = request.json
        day = data.get('research_day', '').lower()
        if day not in ('tuesday', 'friday'):
            return jsonify({'error': 'Invalid value. Must be "tuesday" or "friday"'}), 400

        set_setting('research_day', day)
        return jsonify({'success': True, 'research_day': day}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/settings/show-fire-crystals', methods=['GET'])
def get_fire_crystals_setting():
    """Get whether fire crystal fields should be shown (public endpoint)."""
    return jsonify({'show_fire_crystals': get_show_fire_crystals()}), 200

@app.route('/api/admin/settings/show-fire-crystals', methods=['PUT'])
def set_fire_crystals_setting():
    """Toggle fire crystal fields visibility."""
    try:
        auth_error = check_admin_auth()
        if auth_error:
            return auth_error

        data = request.json
        show = data.get('show_fire_crystals', False)
        set_setting('show_fire_crystals', 'true' if show else 'false')
        return jsonify({'success': True, 'show_fire_crystals': show}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/settings/time-slot-offset', methods=['GET'])
def get_time_slot_offset_setting():
    """Get the slot offset (minutes) the app uses to generate 30-min slots."""
    return jsonify({
        'time_slot_offset': get_time_slot_offset(),
        'valid_offsets': list(VALID_SLOT_OFFSETS),
    }), 200


@app.route('/api/admin/settings/time-slot-offset', methods=['PUT'])
def set_time_slot_offset_setting():
    """Change the slot offset. Allowed values: -20, -15, -10, 0."""
    try:
        auth_error = check_admin_auth()
        if auth_error:
            return auth_error

        data = request.json or {}
        if 'time_slot_offset' not in data:
            return jsonify({'error': 'time_slot_offset is required'}), 400

        try:
            offset = set_time_slot_offset(data['time_slot_offset'])
        except ValueError as exc:
            return jsonify({'error': str(exc)}), 400

        return jsonify({'success': True, 'time_slot_offset': offset}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/settings/application-closing-time', methods=['PUT'])
def set_closing_time_setting():
    """Set or clear application closing time."""
    try:
        auth_error = check_admin_auth()
        if auth_error:
            return auth_error

        data = request.json
        closing_time = data.get('closing_time', '')

        # Validate if non-empty
        if closing_time:
            try:
                datetime.fromisoformat(closing_time.replace('Z', '+00:00'))
            except ValueError:
                return jsonify({'error': 'Invalid datetime format'}), 400

        set_setting('application_closing_time', closing_time)
        return jsonify({'success': True, 'closing_time': closing_time}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/settings/state-number', methods=['PUT'])
def set_state_number_setting():
    """Set the state number for the welcome message."""
    try:
        auth_error = check_admin_auth()
        if auth_error:
            return auth_error

        data = request.json
        state_number = str(data.get('state_number', '')).strip()
        if not state_number:
            return jsonify({'error': 'State number is required'}), 400

        set_setting('state_number', state_number)
        return jsonify({'success': True, 'state_number': state_number}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/login', methods=['POST'])
def admin_login():
    """Authenticate admin or minister user."""
    try:
        data = request.json
        password = data.get('password', '')

        if password == ADMIN_PASSWORD:
            logger.info("Admin login successful")
            return jsonify({
                'success': True,
                'role': 'admin',
                'token': 'admin-token'
            }), 200
        elif password == MINISTER_PASSWORD:
            logger.info("Minister login successful")
            return jsonify({
                'success': True,
                'role': 'minister',
                'token': 'minister-token'
            }), 200
        else:
            logger.warning("Failed login attempt")
            return jsonify({'error': 'Invalid password'}), 401

    except Exception as e:
        logger.error(f"Error in admin_login: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/players', methods=['GET'])
def get_players():
    """Get all players with calculated points."""
    try:
        # Simple auth check
        auth_error = check_admin_auth()
        if auth_error:
            return auth_error

        players = get_all_players()

        # Add calculated points for each day
        research_day = get_research_day()
        for player in players:
            player['monday_points'] = calculate_points(player, 'monday')
            player['research_points'] = calculate_points(player, research_day)
            player['thursday_points'] = calculate_points(player, 'thursday')
            player['research_day'] = research_day

        return jsonify(players), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/player/<int:player_id>', methods=['PUT'])
def update_player(player_id):
    """Update player information."""
    try:
        # Simple auth check
        auth_error = check_admin_auth()
        if auth_error:
            return auth_error

        data = request.json
        time_slots_by_day = data.pop('time_slots_by_day', None)
        time_slots = data.pop('time_slots', [])
        if time_slots_by_day:
            time_slots = time_slots_by_day

        # Update player
        save_player(data, time_slots)

        return jsonify({'success': True, 'message': 'Player updated'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/player/<int:player_id>', methods=['DELETE'])
def remove_player(player_id):
    """Delete a player."""
    try:
        # Simple auth check
        auth_error = check_admin_auth()
        if auth_error:
            return auth_error

        delete_player(player_id)
        logger.info(f"Player deleted: id={player_id}")
        return jsonify({'success': True, 'message': 'Player deleted'}), 200
    except Exception as e:
        logger.error(f"Error deleting player {player_id}: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/assignments/auto-assign', methods=['POST'])
def auto_assign():
    """Auto-assign players to time slots based on points."""
    try:
        # Simple auth check
        auth_error = check_admin_auth()
        if auth_error:
            return auth_error

        data = request.json
        day = data.get('day', '').lower()

        research_day = get_research_day()
        valid_days = ['monday', research_day, 'thursday']
        if day not in valid_days:
            return jsonify({'error': 'Invalid day'}), 400

        players = get_all_players()

        # Map day to day_type for time preferences
        day_type_map = {'monday': 'construction', 'thursday': 'troop'}
        # Research day (tuesday or friday) maps to 'research'
        day_type_map[research_day] = 'research'
        day_type = day_type_map.get(day, 'construction')

        # Calculate points for this day
        for player in players:
            player['points'] = calculate_points(player, day)

        # Sort by points (descending)
        players.sort(key=lambda p: p['points'], reverse=True)

        # Generate the slot grid for the configured offset.
        # Default offset is -10, which yields the same 49-slot end-of-day-anchored
        # layout the app used to hardcode: ['23:50', '00:20', ..., '23:50+'].
        slot_offset = get_time_slot_offset()
        time_slots = slot_ids(slot_offset)

        # Fetch sticky assignments BEFORE clearing
        db = get_db()
        cursor = db.cursor()
        cursor.execute('''
            SELECT a.player_id, a.slot_index, p.fid, p.game_name,
                   p.avatar_image, p.stove_lv, p.stove_lv_content, p.alliance,
                   p.construction_speedups_days, p.research_speedups_days,
                   p.troop_training_speedups_days, p.general_speedups_days,
                   p.fire_crystals, p.refined_fire_crystals, p.fire_crystal_shards
            FROM assignments a
            JOIN players p ON a.player_id = p.id
            WHERE a.day = ? AND a.is_sticky = 1
        ''', (day,))
        sticky_rows = cursor.fetchall()
        sticky_slots = {}  # slot_id (string) -> player data
        sticky_player_ids = set()
        for row in sticky_rows:
            row_dict = dict(row)
            row_dict['points'] = calculate_points(row_dict, day)
            slot_str = slot_index_to_id(row['slot_index'], slot_offset)
            if slot_str is None:
                # Orphaned sticky row — slot_index doesn't map under the
                # currently configured offset (for example offset was lowered
                # from -10 to 0 and a slot_index=48 row was left behind).
                # Skip rather than 500; the row stays in storage so the admin
                # can recover by reverting the offset.
                logger.warning(
                    "Skipping sticky assignment for player_id=%s day=%s: "
                    "slot_index=%s is out of range for offset=%s",
                    row['player_id'], day, row['slot_index'], slot_offset,
                )
                continue
            sticky_slots[slot_str] = {
                'id': row['player_id'],
                'player_id': row['player_id'],
                'fid': row['fid'],
                'game_name': row['game_name'],
                'points': row_dict['points'],
                'avatar_image': row_dict.get('avatar_image') or '',
                'stove_lv': row_dict.get('stove_lv'),
                'stove_lv_content': row_dict.get('stove_lv_content') or '',
                'alliance': row_dict.get('alliance') or '',
                'is_sticky': True,
            }
            sticky_player_ids.add(row['player_id'])

        # Assignment logic
        assignments = {slot: [] for slot in time_slots}
        unassigned = []

        # Pre-fill sticky assignments
        for slot, player_data in sticky_slots.items():
            if slot in assignments:
                assignments[slot].append(player_data)

        for player in players:
            # Skip players that are sticky-assigned
            if player['id'] in sticky_player_ids:
                continue

            # Use day-specific time preferences
            time_slots_by_day = player.get('time_slots_by_day', {})
            player_time_prefs = set(time_slots_by_day.get(day_type, player.get('time_slots', [])))

            # Find matching 30-min slots for player's hourly preferences.
            # For non-zero offsets this is 3 slots per hour (±20 min tolerance);
            # for offset 0 it is the two aligned half-hour slots.
            matching_slots = []
            for pref in player_time_prefs:
                if ':' in pref:
                    h = int(pref.split(':')[0])
                    matching_slots.extend(matching_slots_for_hour(h, slot_offset))

            assigned = False
            for slot in matching_slots:
                if slot in assignments and len(assignments[slot]) == 0:
                    assignments[slot].append({
                        'id': player['id'],
                        'player_id': player['id'],
                        'fid': player['fid'],
                        'game_name': player['game_name'],
                        'points': player['points'],
                        'avatar_image': player.get('avatar_image', ''),
                        'stove_lv': player.get('stove_lv'),
                        'stove_lv_content': player.get('stove_lv_content', ''),
                        'alliance': player.get('alliance', ''),
                        'is_sticky': False,
                    })
                    assigned = True
                    break

            if not assigned:
                unassigned.append({
                    'id': player['id'],
                    'player_id': player['id'],
                    'fid': player['fid'],
                    'game_name': player['game_name'],
                    'points': player['points'],
                    'preferred_times': list(player_time_prefs),
                    'avatar_image': player.get('avatar_image', ''),
                    'stove_lv': player.get('stove_lv'),
                    'stove_lv_content': player.get('stove_lv_content', ''),
                    'alliance': player.get('alliance', ''),
                    'is_sticky': False,
                })

        # Clear existing non-sticky assignments for this day, then clear sticky too (we'll re-insert all)
        cursor.execute('DELETE FROM assignments WHERE day = ?', (day,))

        # Save new assignments (including sticky ones). Convert slot ID strings
        # back to numerical indices on write so the schema is offset-independent.
        for time_slot, slot_players in assignments.items():
            slot_idx = slot_id_to_index(time_slot, slot_offset)
            if slot_idx is None:
                continue
            for position, player in enumerate(slot_players):
                cursor.execute('''
                    INSERT INTO assignments (player_id, day, slot_index, position, is_assigned, is_sticky)
                    VALUES (?, ?, ?, ?, 1, ?)
                ''', (player['id'], day, slot_idx, position, 1 if player.get('is_sticky') else 0))

        db.commit()

        assigned_count = sum(1 for s in assignments.values() if s)
        logger.info(f"Auto-assign completed: day={day}, assigned={assigned_count}, unassigned={len(unassigned)}")

        return jsonify({
            'success': True,
            'assignments': assignments,
            'unassigned': unassigned,
            'slot_mapping': slot_mapping(slot_offset),
        }), 200

    except Exception as e:
        logger.error(f"Error in auto_assign: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/assignments/<day>', methods=['GET'])
def get_assignments(day):
    """Get assignments for a specific day."""
    try:
        # Simple auth check
        auth_error = check_admin_auth()
        if auth_error:
            return auth_error

        slot_offset = get_time_slot_offset()
        db = get_db()
        cursor = db.cursor()
        cursor.execute('''
            SELECT
                a.id, a.slot_index, a.position, a.is_assigned, a.is_sticky,
                p.id as player_id, p.fid, p.game_name,
                p.construction_speedups_days, p.research_speedups_days,
                p.troop_training_speedups_days, p.general_speedups_days,
                p.fire_crystals, p.refined_fire_crystals, p.fire_crystal_shards,
                p.avatar_image, p.stove_lv, p.stove_lv_content, p.alliance
            FROM assignments a
            JOIN players p ON a.player_id = p.id
            WHERE a.day = ?
            ORDER BY a.slot_index, a.position
        ''', (day.lower(),))

        rows = cursor.fetchall()
        assignments = {}
        for row in rows:
            slot = slot_index_to_id(row['slot_index'], slot_offset)
            if slot is None:
                continue
            if slot not in assignments:
                assignments[slot] = []
            row_dict = dict(row)
            row_dict['time_slot'] = slot  # back-compat for clients reading the field
            row_dict['points'] = calculate_points(row_dict, day.lower())
            assignments[slot].append(row_dict)

        return jsonify(assignments), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/assignments/update', methods=['POST'])
def update_assignments():
    """Update assignments after drag-and-drop."""
    try:
        # Simple auth check
        auth_error = check_admin_auth()
        if auth_error:
            return auth_error

        data = request.json
        day = data.get('day')
        assignments = data.get('assignments', {})
        slot_offset = get_time_slot_offset()

        slot_indices_by_id = {}
        invalid_slots = []
        for time_slot, slot_players in assignments.items():
            if not slot_players:
                continue
            slot_idx = slot_id_to_index(time_slot, slot_offset)
            if slot_idx is None:
                invalid_slots.append(time_slot)
            else:
                slot_indices_by_id[time_slot] = slot_idx

        if invalid_slots:
            logger.warning(
                "Rejected assignment update for day=%s at offset=%s due to invalid slot IDs: %s",
                day, slot_offset, invalid_slots,
            )
            return jsonify({
                'error': 'Invalid time slot(s) for current offset',
                'invalid_slots': invalid_slots,
                'time_slot_offset': slot_offset,
            }), 400

        db = get_db()
        cursor = db.cursor()

        # Clear existing assignments
        cursor.execute('DELETE FROM assignments WHERE day = ?', (day,))

        # Save new assignments (enforce max 1 player per slot)
        for time_slot, slot_players in assignments.items():
            if not slot_players:
                continue
            slot_idx = slot_indices_by_id[time_slot]
            player = slot_players[0]  # Only take first player per slot
            cursor.execute('''
                INSERT INTO assignments (player_id, day, slot_index, position, is_assigned, is_sticky)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (player['player_id'], day, slot_idx, 0,
                  player.get('is_assigned', True),
                  1 if player.get('is_sticky') else 0))

        db.commit()

        return jsonify({'success': True}), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/export', methods=['GET'])
def export_assignments():
    """Export assignments for all days to a single Excel workbook."""
    try:
        # Simple auth check
        auth_error = check_admin_auth()
        if auth_error:
            return auth_error

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        separator_fill = PatternFill(start_color="F4B942", end_color="F4B942", fill_type="solid")
        separator_font = Font(bold=True, color="000000")

        slot_offset = get_time_slot_offset()

        headers = [
            'Time Slot', 'FID', 'Alliance', 'Game Name',
            'Construction (days)', 'Research (days)',
            'Troop Training (days)', 'General (days)',
            'Fire Crystals', 'Refined Fire Crystals',
            'Crystal Shards', 'Points'
        ]

        col_widths = [15, 15, 10, 25, 18, 15, 20, 15, 13, 18, 14, 12]

        research_day = get_research_day()
        research_label = 'Tuesday - Research' if research_day == 'tuesday' else 'Friday - Research'
        days = [
            ('monday', 'Monday - Construction'),
            (research_day, research_label),
            ('thursday', 'Thursday - Troop Training'),
        ]

        db = get_db()
        cursor = db.cursor()

        for day_key, day_title in days:
            ws = wb.create_sheet(title=day_title)
            ws.append(headers)

            # Style headers
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            # Get assigned players
            cursor.execute('''
                SELECT a.slot_index, p.*
                FROM assignments a
                JOIN players p ON a.player_id = p.id
                WHERE a.day = ? AND a.is_assigned = 1
                ORDER BY a.slot_index, a.position
            ''', (day_key,))

            assigned_rows = cursor.fetchall()
            assigned_player_ids = set()

            for row in assigned_rows:
                player = dict(row)
                assigned_player_ids.add(player['id'])
                points = calculate_points(player, day_key)
                slot_str = slot_index_to_id(row['slot_index'], slot_offset) or ''
                ws.append([
                    display_slot_id(slot_str),
                    player['fid'],
                    player.get('alliance', ''),
                    player['game_name'],
                    player['construction_speedups_days'],
                    player['research_speedups_days'],
                    player['troop_training_speedups_days'],
                    player['general_speedups_days'],
                    player['fire_crystals'],
                    player['refined_fire_crystals'],
                    player['fire_crystal_shards'],
                    points,
                ])

            # Get all players to find unassigned ones
            cursor.execute('SELECT * FROM players ORDER BY id')
            all_players = [dict(r) for r in cursor.fetchall()]
            unassigned = [p for p in all_players if p['id'] not in assigned_player_ids]

            # Sort unassigned by points descending
            for p in unassigned:
                p['_points'] = calculate_points(p, day_key)
            unassigned.sort(key=lambda p: p['_points'], reverse=True)

            if unassigned:
                # Separator row
                sep_row_num = ws.max_row + 2  # skip a blank row
                ws.append([])  # blank row
                ws.append(['UNASSIGNED PLAYERS'] + [''] * (len(headers) - 1))
                for cell in ws[sep_row_num]:
                    cell.fill = separator_fill
                    cell.font = separator_font

                for player in unassigned:
                    ws.append([
                        'Unassigned',
                        player['fid'],
                        player.get('alliance', ''),
                        player['game_name'],
                        player['construction_speedups_days'],
                        player['research_speedups_days'],
                        player['troop_training_speedups_days'],
                        player['general_speedups_days'],
                        player['fire_crystals'],
                        player['refined_fire_crystals'],
                        player['fire_crystal_shards'],
                        player['_points'],
                    ])

            # Set column widths
            for i, width in enumerate(col_widths):
                ws.column_dimensions[chr(65 + i)].width = width

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f'ministry_assignments_{datetime.now().strftime("%Y%m%d")}.xlsx'
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/settings/publish', methods=['PUT'])
def publish_schedule():
    """Publish assignments for a day so they appear on the public page.
    Supports multiple days — stores as comma-separated list.
    """
    try:
        auth_error = check_admin_auth()
        if auth_error:
            return auth_error

        data = request.json
        day = data.get('day', '').lower()
        research_day = get_research_day()
        valid_days = ['monday', research_day, 'thursday']
        if day not in valid_days:
            return jsonify({'error': 'Invalid day'}), 400

        # Add to existing published days
        current = get_setting('published_days', '')
        days_set = set(d for d in current.split(',') if d)
        days_set.add(day)
        set_setting('published_days', ','.join(sorted(days_set)))
        return jsonify({'success': True, 'published_days': sorted(days_set)}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/settings/unpublish', methods=['PUT'])
def unpublish_schedule():
    """Remove a specific day from published schedules."""
    try:
        auth_error = check_admin_auth()
        if auth_error:
            return auth_error

        data = request.json
        day = data.get('day', '').lower()

        current = get_setting('published_days', '')
        days_set = set(d for d in current.split(',') if d)
        days_set.discard(day)
        set_setting('published_days', ','.join(sorted(days_set)))
        return jsonify({'success': True, 'published_days': sorted(days_set)}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/published-schedule/<day>', methods=['GET'])
def get_published_schedule(day):
    """Get the published schedule for a specific day (public endpoint).
    Returns only player name, alliance, and time slot — no points or resources.
    """
    try:
        published_days = get_setting('published_days', '')
        days_list = [d for d in published_days.split(',') if d]
        if day.lower() not in days_list:
            return jsonify({'published': False}), 200

        slot_offset = get_time_slot_offset()
        db = get_db()
        cursor = db.cursor()
        cursor.execute('''
            SELECT a.slot_index, p.game_name, p.alliance
            FROM assignments a
            JOIN players p ON a.player_id = p.id
            WHERE a.day = ? AND a.is_assigned = 1
            ORDER BY a.slot_index, a.position
        ''', (day.lower(),))

        rows = cursor.fetchall()
        assignments = {}
        for row in rows:
            slot = slot_index_to_id(row['slot_index'], slot_offset)
            if slot is None:
                continue
            if slot not in assignments:
                assignments[slot] = []
            assignments[slot].append({
                'game_name': row['game_name'],
                'alliance': row['alliance'] or '',
            })

        day_labels = {
            'monday': 'Monday - Construction',
            'tuesday': 'Tuesday - Research',
            'friday': 'Friday - Research',
            'thursday': 'Thursday - Troop Training',
        }

        return jsonify({
            'published': True,
            'day': day.lower(),
            'day_label': day_labels.get(day.lower(), day),
            'assignments': assignments,
            'slot_mapping': slot_mapping(slot_offset),
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/settings/published-days', methods=['GET'])
def get_published_days_setting():
    """Get which days are currently published (public endpoint)."""
    published_days = get_setting('published_days', '')
    days_list = [d for d in published_days.split(',') if d]
    return jsonify({'published_days': days_list}), 200


@app.route('/api/settings/application-closing-time', methods=['GET'])
def get_closing_time_setting():
    """Get application closing time and whether applications are currently closed."""
    closing_time = get_setting('application_closing_time', '')
    is_closed = False
    if closing_time:
        try:
            dt = datetime.fromisoformat(closing_time.replace('Z', '+00:00'))
            is_closed = datetime.now(timezone.utc) >= dt
        except ValueError:
            pass
    return jsonify({'closing_time': closing_time, 'is_closed': is_closed}), 200


@app.route('/api/settings/state-number', methods=['GET'])
def get_state_number_setting():
    """Get the state number for the welcome message."""
    return jsonify({'state_number': get_setting('state_number', '2694')}), 200


@app.route('/api/time-preferences/heatmap', methods=['GET'])
def get_heatmap():
    """Get time preference counts per slot per day_type (public endpoint)."""
    try:
        counts = get_time_preference_counts()
        return jsonify(counts), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/player/<fid>/assignments', methods=['GET'])
def get_player_assignments(fid):
    """Get a player's current assignments by FID (public endpoint).
    Returns only day and time_slot — no points or resource data.
    """
    try:
        player = get_player_by_fid(fid)
        if not player:
            return jsonify({'error': 'Player not found'}), 404

        slot_offset = get_time_slot_offset()
        db = get_db()
        cursor = db.cursor()
        cursor.execute('''
            SELECT day, slot_index
            FROM assignments
            WHERE player_id = ? AND is_assigned = 1
            ORDER BY day, slot_index
        ''', (player['id'],))

        assignments = {}
        for row in cursor.fetchall():
            day = row['day']
            slot = slot_index_to_id(row['slot_index'], slot_offset)
            if slot is None:
                continue
            if day not in assignments:
                assignments[day] = []
            assignments[day].append({'time_slot': slot})

        return jsonify({
            'assignments': assignments,
            'slot_mapping': slot_mapping(slot_offset),
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/players/export-json', methods=['GET'])
def export_players_json():
    """Export all players as JSON (admin auth required)."""
    try:
        auth_error = check_admin_auth()
        if auth_error:
            return auth_error

        players = get_all_players()

        # Clean up internal fields for export
        export_players = []
        for p in players:
            export_players.append({
                'fid': p['fid'],
                'game_name': p['game_name'],
                'alliance': p.get('alliance', ''),
                'construction_speedups_days': p['construction_speedups_days'],
                'research_speedups_days': p['research_speedups_days'],
                'troop_training_speedups_days': p['troop_training_speedups_days'],
                'general_speedups_days': p['general_speedups_days'],
                'fire_crystals': p['fire_crystals'],
                'refined_fire_crystals': p['refined_fire_crystals'],
                'fire_crystal_shards': p['fire_crystal_shards'],
                'avatar_image': p.get('avatar_image', ''),
                'stove_lv': p.get('stove_lv'),
                'stove_lv_content': p.get('stove_lv_content', ''),
                'timezone': p.get('timezone', ''),
                'time_slots_by_day': p.get('time_slots_by_day', {}),
            })

        export_data = {
            'version': 1,
            'exported_at': datetime.now().isoformat(),
            'players': export_players,
        }

        output = json.dumps(export_data, indent=2)
        filename = f'players_backup_{datetime.now().strftime("%Y%m%d")}.json'
        return Response(
            output,
            mimetype='application/json',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/players/import', methods=['POST'])
def import_players_json():
    """Import players from JSON (admin auth required). Upserts by FID."""
    try:
        auth_error = check_admin_auth()
        if auth_error:
            return auth_error

        data = request.json
        if not data or 'players' not in data:
            return jsonify({'error': 'Invalid format: expected {players: [...]}'}), 400

        imported = 0
        updated = 0
        errors = 0

        for p in data['players']:
            try:
                fid = p.get('fid', '').strip()
                if not fid:
                    errors += 1
                    continue

                # Check if player exists
                existing = get_player_by_fid(fid)

                player_data = {
                    'fid': fid,
                    'game_name': p.get('game_name', 'Unknown'),
                    'alliance': p.get('alliance', ''),
                    'construction_speedups_days': float(p.get('construction_speedups_days', 0)),
                    'research_speedups_days': float(p.get('research_speedups_days', 0)),
                    'troop_training_speedups_days': float(p.get('troop_training_speedups_days', 0)),
                    'general_speedups_days': float(p.get('general_speedups_days', 0)),
                    'fire_crystals': int(p.get('fire_crystals', 0)),
                    'refined_fire_crystals': int(p.get('refined_fire_crystals', 0)),
                    'fire_crystal_shards': int(p.get('fire_crystal_shards', 0)),
                    'avatar_image': p.get('avatar_image', ''),
                    'stove_lv': p.get('stove_lv'),
                    'stove_lv_content': p.get('stove_lv_content', ''),
                    'timezone': p.get('timezone', ''),
                }

                time_slots = p.get('time_slots_by_day', p.get('time_slots', []))
                save_player(player_data, time_slots)

                if existing:
                    updated += 1
                else:
                    imported += 1
            except Exception:
                errors += 1

        return jsonify({
            'success': True,
            'imported': imported,
            'updated': updated,
            'errors': errors,
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/players/delete-all', methods=['DELETE'])
def delete_all_players():
    """Delete all players and their assignments."""
    try:
        auth_error = check_admin_auth()
        if auth_error:
            return auth_error

        db = get_db()
        cursor = db.cursor()
        cursor.execute('DELETE FROM assignments')
        cursor.execute('DELETE FROM time_preferences')
        cursor.execute('DELETE FROM players')
        db.commit()

        logger.warning("All players deleted by admin")
        return jsonify({'success': True, 'message': 'All players deleted'}), 200
    except Exception as e:
        logger.error(f"Error deleting all players: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500

# Optional URL sub-path prefix (e.g. "/ministry"). When set, the Flask app is
# mounted under that prefix so it can run behind a path-based reverse proxy
# (e.g. Cloudflare Tunnel routing /ministry/* to this service). /health is
# kept at the root so platform health probes don't need to be prefix-aware.
URL_PREFIX = os.getenv('URL_PREFIX', '').rstrip('/')
if URL_PREFIX:
    from werkzeug.middleware.dispatcher import DispatcherMiddleware
    from werkzeug.wrappers import Response as WzResponse

    def _root_app(environ, start_response):
        if environ.get('PATH_INFO') == '/health':
            return WzResponse('{"status":"healthy"}', mimetype='application/json')(environ, start_response)
        return WzResponse('Not Found', status=404)(environ, start_response)

    app.config['APPLICATION_ROOT'] = URL_PREFIX
    app.config['SESSION_COOKIE_PATH'] = URL_PREFIX
    app.wsgi_app = DispatcherMiddleware(_root_app, {URL_PREFIX: app.wsgi_app})

if __name__ == '__main__':
    port = int(os.getenv('PORT', 8080))
    app.run(host='0.0.0.0', port=port, debug=os.getenv('FLASK_ENV') == 'development')
